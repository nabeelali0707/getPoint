import re
import sys
import os
import json
import random
import openpyxl
from pypdf import PdfReader

# FAST NU Karachi Campus Coordinate
FAST_NU_COORDS = (24.8569, 67.2647)

KNOWN_COORDS = {
    "fast nu": FAST_NU_COORDS,
    "fast": FAST_NU_COORDS,
    "fast campus": FAST_NU_COORDS,
    "malir cantt": (24.9082, 67.2023),
    "askari-v": (24.9082, 67.2023),
    "malir cantt Askari-v": (24.9082, 67.2023),
    "safoora": (24.9195, 67.1485),
    "disco bakery": (24.9298, 67.0984),
    "maskan": (24.9312, 67.0970),
    "hassan square": (24.9036, 67.0781),
    "soldier bazar": (24.8722, 67.0264),
    "teen talwar": (24.8340, 67.0328),
    "bahria town": (24.9880, 67.4520),
    "tower": (24.8485, 67.0012),
    "kpt bridge": (24.8252, 67.0543),
    "ftc": (24.8540, 67.0505),
    "gulshan": (24.9180, 67.0970),
    "maymar": (25.0112, 67.1264),
    "korangi": (24.8200, 67.1200),
    "johar": (24.9100, 67.1200),
    "g3 stop": (24.9255, 67.1430),
    "pahalwan goath": (24.9080, 67.1350),
    "rimjim hall": (24.9280, 67.0920),
    "power house": (24.9750, 67.0680),
}

def parse_phone(text):
    if not text:
        return None
    match = re.search(r'(03\d{2}[-\s]?\d{7})', text)
    if match:
        return match.group(1).replace(" ", "").replace("-", "")
    return None

def parse_driver_name(text, phone):
    if not text:
        return None
    if phone:
        cleaned = text
        for p_part in [phone, phone[:4], phone[4:], phone[:4]+"-"+phone[4:]]:
            cleaned = cleaned.replace(p_part, "")
        text = cleaned
    text = text.replace("Name:", "").replace("Cell #", "").replace("Cell", "").strip()
    text = re.sub(r'[-\s:#]+$', '', text).strip()
    return text

def get_base_coords(stop_name):
    name_lower = stop_name.lower()
    for key, coords in KNOWN_COORDS.items():
        if key in name_lower:
            return coords
    lat = random.uniform(24.83, 24.96)
    lng = random.uniform(67.00, 67.18)
    return (lat, lng)

def geocode_stops(stops):
    if not stops:
        return []
    stops = sorted(stops, key=lambda x: x["sequence"])
    n = len(stops)
    if n == 0:
        return []
    start_coords = get_base_coords(stops[0]["name"])
    end_coords = FAST_NU_COORDS
    
    geocoded = []
    for i, stop in enumerate(stops):
        if i == 0:
            lat, lng = start_coords
        elif i == n - 1:
            lat, lng = end_coords
        else:
            ratio = i / (n - 1)
            lat = start_coords[0] + ratio * (end_coords[0] - start_coords[0])
            lng = start_coords[1] + ratio * (end_coords[1] - start_coords[1])
            lat += random.uniform(-0.0015, 0.0015)
            lng += random.uniform(-0.0015, 0.0015)
        
        geocoded.append({
            "name": stop["name"],
            "sequence": stop["sequence"],
            "scheduledTime": stop["scheduledTime"],
            "lat": round(lat, 6),
            "lng": round(lng, 6)
        })
    return geocoded

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "No file path provided"}))
        return

    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(json.dumps({"error": "File does not exist"}))
        return

    routes = []
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsx":
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        current_route = None
        current_stops = []
        
        for r in range(1, sheet.max_row + 1):
            cell_1 = sheet.cell(row=r, column=1).value
            cell_2 = sheet.cell(row=r, column=2).value
            
            if cell_1 and isinstance(cell_1, str) and ("Jadoon R:" in cell_1 or "R:" in cell_1):
                if current_route and current_stops:
                    current_route["stops"] = geocode_stops(current_stops)
                    current_route["startTime"] = current_stops[0]["scheduledTime"]
                    current_route["endTime"] = current_stops[-1]["scheduledTime"]
                    routes.append(current_route)
                
                code = cell_1.strip()
                current_route = {
                    "code": code,
                    "provider": "Jadoon",
                    "driverName": None,
                    "driverPhone": None,
                    "startTime": "06:30",
                    "endTime": "07:45",
                    "stops": []
                }
                current_stops = []
            
            elif current_route and cell_1 and cell_2 == "Time":
                phone = parse_phone(cell_1)
                name = parse_driver_name(cell_1, phone)
                current_route["driverName"] = name
                current_route["driverPhone"] = phone
                
            elif current_route and cell_1 and cell_2:
                stop_name = str(cell_1).strip()
                if isinstance(cell_2, str):
                    time_str = cell_2.strip()
                elif hasattr(cell_2, "strftime"):
                    time_str = cell_2.strftime("%H:%M")
                else:
                    time_str = str(cell_2).strip()
                
                current_stops.append({
                    "name": stop_name,
                    "sequence": len(current_stops),
                    "scheduledTime": time_str
                })
        
        if current_route and current_stops:
            current_route["stops"] = geocode_stops(current_stops)
            current_route["startTime"] = current_stops[0]["scheduledTime"]
            current_route["endTime"] = current_stops[-1]["scheduledTime"]
            routes.append(current_route)

    elif ext == ".pdf":
        reader = PdfReader(file_path)
        for p_idx, page in enumerate(reader.pages):
            text = page.extract_text()
            lines = [l.strip() for l in text.split("\n") if l.strip()]
            
            route_code = f"Nadeem R:{p_idx+1}"
            driver_name = None
            driver_phone = None
            stops_raw = []
            
            for line in lines[:5]:
                if "ROUTE #" in line:
                    route_code = line.replace("ROUTE #", "Nadeem R:").strip()
                    break
            
            for line in lines[:8]:
                if "Name:" in line or "Cell" in line:
                    driver_phone = parse_phone(line)
                    driver_name = parse_driver_name(line, driver_phone)
                    break
            
            is_timing = False
            temp_stop_name = ""
            
            for line in lines:
                if "ROUTE TIMING" in line or "TIMING" in line:
                    is_timing = True
                    continue
                
                if is_timing:
                    time_match = re.search(r'(\d{1,2}:\d{2}\s*(?:am|pm|AM|PM)?)$', line)
                    if time_match:
                        time_str = time_match.group(1)
                        name_part = line.replace(time_str, "").strip()
                        
                        if temp_stop_name:
                            name_part = temp_stop_name + " " + name_part
                            temp_stop_name = ""
                            
                        stops_raw.append({
                            "name": name_part,
                            "sequence": len(stops_raw),
                            "scheduledTime": time_str
                        })
                    else:
                        if not any(header in line for header in ["FALL", "ROUTE #", "Name:", "Cell #"]):
                            temp_stop_name = (temp_stop_name + " " + line).strip()
            
            if stops_raw:
                routes.append({
                    "code": route_code,
                    "provider": "Nadeem",
                    "driverName": driver_name,
                    "driverPhone": driver_phone,
                    "startTime": stops_raw[0]["scheduledTime"],
                    "endTime": stops_raw[-1]["scheduledTime"],
                    "stops": geocode_stops(stops_raw)
                })

    print(json.dumps(routes))

if __name__ == "__main__":
    main()
